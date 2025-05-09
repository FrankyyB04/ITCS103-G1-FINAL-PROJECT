import tkinter as tk
from tkinter import messagebox, simpledialog, StringVar, ttk, Button
import os
import datetime
from openpyxl import Workbook, load_workbook

# ------------------- Constants -------------------

EXCEL_FILE = "coffee_shop_data.xlsx"
ADMIN_CREDENTIALS = {"username": "admin", "password": "admin123"}

# ------------------- Globals -------------------

hot_coffee_items = []
snacks_items = []
iced_tea_items = []
iced_coffee_items = []

# ------------------- Excel Handling -------------------
def ensure_excel_file():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws_users = wb.active
        ws_users.title = "Users"
        ws_users.append(["Username", "Password"])
        wb.create_sheet(title="Menu").append(["Section", "Name", "Price", "Stock"])
        wb.create_sheet(title="Orders").append(["Order Number", "Order Date", "Item Name", "Quantity", "Price", "Total"])
        wb.save(EXCEL_FILE)
def load_users_from_excel():
    ensure_excel_file()
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Users"]
    return {row[0]: row[1] for row in ws.iter_rows(min_row=2, values_only=True) if row[0] and row[1]}
def add_user_to_excel(username, password):
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Users"]
    ws.append([username, password])
    wb.save(EXCEL_FILE)
def load_menu_from_excel():
    ensure_excel_file()
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Menu"]
    hot, snacks, tea, coffee = [], [], [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        section, name, price, stock = row if len(row) == 4 else (*row, 0)
        item = (name, price, stock)
        if section == "hot":
            hot.append(item)
        elif section == "snacks":
            snacks.append(item)
        elif section == "iced_tea":
            tea.append(item)
        elif section == "iced_coffee":
            coffee.append(item)
    return hot, snacks, tea, coffee
def save_menu_to_excel():
    wb = load_workbook(EXCEL_FILE)
    if "Menu" in wb.sheetnames:
        del wb["Menu"]
    ws = wb.create_sheet("Menu")
    ws.append(["Section", "Name", "Price", "Stock"])
    for section, items in [("hot", hot_coffee_items), ("snacks", snacks_items), ("iced_tea", iced_tea_items), ("iced_coffee", iced_coffee_items)]:
        for name, price, stock in items:
            ws.append([section, name, price, stock])
    wb.save(EXCEL_FILE)    
def save_order_to_excel(order_number, order_date, selected_items, total):
    ensure_excel_file()
    wb = load_workbook(EXCEL_FILE)
    ws_orders = wb["Orders"]
    ws_menu = wb["Menu"]

    # Save the order
    for name, price, quantity in selected_items:
        ws_orders.append([order_number, order_date, name, quantity, price, price * quantity])

    # Deduct stock
    for row in ws_menu.iter_rows(min_row=2):
        item_name = row[1].value
        for name, price, quantity in selected_items:
            if item_name == name:
                try:
                    current_stock = int(row[3].value)
                    if current_stock >= quantity:
                        row[3].value = current_stock - quantity
                    else:
                        row[3].value = 0  # prevent negative stock
                except:
                    continue  # if stock value is missing or invalid, skip
    wb.save(EXCEL_FILE)

# ------------------- Menu Globals -------------------

def load_menu_globals():
    global hot_coffee_items, snacks_items, iced_tea_items, iced_coffee_items
    hot_coffee_items, snacks_items, iced_tea_items, iced_coffee_items = load_menu_from_excel()

# ------------------- Ordering -------------------

def open_ordering_window(username):
    order_window = tk.Toplevel(main_window)
    main_window.withdraw()
    load_menu_globals()
    order_window.geometry("400x500")
    order_window.resizable(False, False)
    order_window.configure(bg='#E0CAA0')
    order_window.title(f"{username}'s Order")

    
    top_frame = tk.Frame(order_window, bg='#E0CAA0')
    top_frame.pack(side='top', fill='x')

    
    back_button = tk.Button(top_frame, text="<", font=("lets coffee", 10, 'bold'),
                            bg='#2B1E15', fg='#E0CAA0', relief='flat',
                            command=lambda: [order_window.destroy(), open_post_login_menu(username)])
    back_button.pack(side='left', padx=10, pady=10)

    
    title_label = tk.Label(top_frame, text="Select Products", font=('Lets coffee', 15, 'bold'),
                           bg='#E0CAA0', fg='#2B1E15')
    title_label.pack(side='top', pady=10,padx=0)

    
    canvas_frame = tk.Frame(order_window, bg='#E0CAA0')
    canvas_frame.pack(fill="both", expand=True)

    canvas = tk.Canvas(canvas_frame, bg='#E0CAA0', highlightthickness=0)
    scrollbar = tk.Scrollbar(canvas_frame, orient="vertical", command=canvas.yview)
    scrollable_frame = tk.Frame(canvas, bg='#E0CAA0')

    scrollable_frame.bind(
        "<Configure>",
        lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
    )

    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)

    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")

    
    vars_list = []

    def add_section(section_name, items, row_offset):
        frame = tk.LabelFrame(scrollable_frame, text=section_name.title(), bg='#E0CAA0')
        frame.grid(row=row_offset, column=0, padx=40, pady=10, sticky='ew')
        row_num = 0
        for name, price, stock in items:
            var = tk.IntVar()
            quantity_var = tk.IntVar(value=1)
            tk.Checkbutton(frame, 
                           text=f"{name}   ---   ₱ {price}", 
                           variable=var,
                           bg='#E0CAA0').grid(row=row_num, column=0, sticky="w", padx=5, pady=2)
            tk.Label(frame, text=f"Available: {stock}", bg='#E0CAA0').grid(row=row_num, column=1, padx=5, sticky="w")
            spinbox = tk.Spinbox(frame, from_=1, to=stock if stock > 0 else 1, textvariable=quantity_var, width=5)
            spinbox.grid(row=row_num, column=2, padx=5, sticky="w")
            if stock == 0:
                spinbox.config(state='disabled')
            vars_list.append((var, name, price, quantity_var, stock))
            row_num += 1
        return row_offset + 1

    row_offset = 0
    for section, items in [("Hot Coffee", hot_coffee_items), 
                           ("Snacks", snacks_items), 
                           ("Iced Tea", iced_tea_items), 
                           ("Iced Coffee", iced_coffee_items)]:
        row_offset = add_section(section, items, row_offset)

    def proceed_to_checkout():
        selected_items = []
        for var, name, price, quantity_var, stock in vars_list:
            if var.get():
                qty = quantity_var.get()
                if qty > stock:
                    return messagebox.showerror("Order Error", f"Not enough stock for {name}.\nPlease adjust the quantity.")
                selected_items.append((name, price, qty))
        if selected_items:
            open_checkout_window(selected_items, username)
            order_window.destroy()
        else:
            messagebox.showwarning("No items", "Please select items to proceed.")

    # === Bottom checkout button (fixed, non-scrollable) ===
    checkout_button = tk.Button(order_window, text="Proceed to Checkout", font=("lets coffee", 12, 'bold'),
                                bg='#2B1E15', fg='#E0CAA0', relief='flat', width=30,
                                command=proceed_to_checkout)
    checkout_button.pack(side="bottom", pady=10)
def get_next_order_number():
    last_order_number = int(open("last_order_number.txt").read().strip()) if os.path.exists("last_order_number.txt") else 0
    next_order_number = last_order_number + 1
    with open("last_order_number.txt", "w") as f:
        f.write(str(next_order_number))
    return f"{next_order_number:03}"
def open_post_login_menu(username):
    menu_win = tk.Toplevel(main_window)
    menu_win.configure(bg='#2B1E15')
    menu_win.geometry('500x500')
    menu_win.title('COFFEE LIFE')

    center_frame = tk.Frame(menu_win, bg='#2B1E15')
    center_frame.pack(expand=True, fill='both', padx=20, pady=20)

    welcome_label = tk.Label(center_frame,
                             text=f"Welcome to Coffee life\n{username}!",
                             font=("cafe coffee", 50, 'bold'),
                             bg='#2B1E15',
                             fg='#E0CAA0')
    welcome_label.pack(pady=30)

    menu_button = tk.Button(center_frame,
                            text="MENU",
                            font=('lets coffee', 15, 'bold'),
                            bg='#E0CAA0',
                            fg='#2B1E15',
                            relief='flat',
                            width=10,
                            command=lambda: [open_ordering_window(username),
                                             menu_win.withdraw()])
    menu_button.pack(pady=10)

    logout_button = tk.Button(center_frame,
                              text="LOG OUT",
                              font=('lets coffee', 15, 'bold'),
                              bg='#E0CAA0',
                              fg='#2B1E15',
                              relief='flat',
                              width=10,
                              command=lambda: [menu_win.destroy(),
                                               username_entry.delete(0, tk.END),
                                               password_entry.delete(0, tk.END),
                                               main_window.deiconify()])
    logout_button.pack(pady=10)

    menu_win.update_idletasks()
    menu_win.minsize(menu_win.winfo_reqwidth(), menu_win.winfo_reqheight())

    main_window.withdraw()
def open_checkout_window(selected_items, username):
    checkout_win = tk.Toplevel(main_window)
    checkout_win.geometry("400x500")
    checkout_win.resizable(False, False)
    checkout_win.title("Checkout")
    checkout_win.configure(bg='#E0CAA0')

    order_number = f"ON-CL-{get_next_order_number()}"
    order_date = datetime.datetime.now().strftime("%m-%d-%Y %H:%M:%S")

    # === Header ===
    header_frame = tk.Frame(checkout_win, bg='#E0CAA0')
    header_frame.pack(pady=10)

    tk.Label(header_frame,
             text="☕ COFFEE LIFE",
             bg='#E0CAA0',
             fg="#2B1E15",
             font=("cafe coffee", 28, 'bold')).pack()

    tk.Label(header_frame,
             text=f"Order Number: {order_number}",
             font=('Arial', 9, 'bold'),
             bg='#E0CAA0').pack(anchor='w', padx=20)

    tk.Label(header_frame,
             text=f"Order Date: {order_date}",
             font=('Arial', 9, 'bold'),
             bg='#E0CAA0').pack(anchor='w', padx=20)

    ttk.Separator(checkout_win, orient='horizontal').pack(fill='x', padx=20, pady=10)

    # === Order Summary ===
    summary_frame = tk.Frame(checkout_win, bg='#F5E7C4', bd=2, relief='ridge')
    summary_frame.pack(padx=20, pady=10, fill='both', expand=True)

    tk.Label(summary_frame,
             text="🛒 Your Order Summary",
             font=('Arial', 10, 'bold'),
             bg='#F5E7C4').pack(anchor='center', pady=10)

    total = 0
    for name, price, quantity in selected_items:
        subtotal = price * quantity
        total += subtotal
        item_text = f"{name}: ₱{price} x {quantity} = ₱{subtotal:.2f}"
        tk.Label(summary_frame,
                 text=item_text,
                 font=('Arial', 9),
                 bg='#F5E7C4').pack(anchor='w', padx=15, pady=2)

    ttk.Separator(summary_frame, orient='horizontal').pack(fill='x', padx=10, pady=10)

    tk.Label(summary_frame,
             text=f"Total: ₱{total:.2f}",
             font=('Arial', 10, 'bold'),
             bg='#F5E7C4',
             fg='#2B1E15').pack(anchor='e', padx=15, pady=5)

    # === Action Buttons ===
    button_frame = tk.Frame(checkout_win, bg='#E0CAA0')
    button_frame.pack(pady=15)

    confirm_btn = tk.Button(button_frame,
                            text=" Confirm Order",
                            font=('arial', 10, 'bold'),
                            bg='#2B1E15',
                            fg='#E0CAA0',
                            relief='flat',
                            width=15,
                            command=lambda: [
                                save_order_to_excel(order_number, order_date, selected_items, total),
                                messagebox.showinfo("Order Confirmed", "Order placed successfully!"),
                                checkout_win.destroy(),
                                open_post_login_menu(username)
                            ])
    confirm_btn.grid(row=0, column=1, padx=10)

    cancel_btn = tk.Button(button_frame,
                           text="Cancel",
                           font=('arial', 10, 'bold'),
                           bg='#2B1E15',
                           fg='#E0CAA0',
                           relief='flat',
                           width=15,
                           command=lambda: [open_ordering_window(username), checkout_win.destroy()])
    cancel_btn.grid(row=0, column=0, padx=10)

# ------------------- Admin -------------------

def open_admin_window():
    admin_window = tk.Toplevel(main_window)
    admin_window.title("Admin Menu Editor")
    admin_window.geometry("900x500")

    # --- Treeview Frame ---
    tree_frame = ttk.Frame(admin_window)
    tree_frame.pack(padx=10, pady=10, fill='both', expand=True)

    tree = ttk.Treeview(tree_frame, columns=("Section", "Name", "Price", "Stock"), show="headings")
    for col in ("Section", "Name", "Price", "Stock"):
        tree.heading(col, text=col)
    tree.pack(fill='both', expand=True)

    for section, items in [("hot", hot_coffee_items), ("snacks", snacks_items), ("iced_tea", iced_tea_items), ("iced_coffee", iced_coffee_items)]:
        for name, price, stock in items:
            tree.insert("", "end", values=(section, name, price, stock))

    # --- Form Frame ---
    form_frame = ttk.Frame(admin_window)
    form_frame.pack(pady=10)

    section_var, name_var, price_var, stock_var = StringVar(), StringVar(), StringVar(), StringVar()
    for i, (label, var) in enumerate([("Section", section_var), ("Name", name_var), ("Price", price_var), ("Stock", stock_var)]):
        ttk.Label(form_frame, text=f"{label}:").grid(row=i, column=0, sticky='e', padx=5, pady=2)
        ttk.Entry(form_frame, textvariable=var).grid(row=i, column=1, padx=5, pady=2)

    # --- Functions ---
    def add_product():
        section, name = section_var.get().strip().lower(), name_var.get().strip()
        try:
            price = float(price_var.get())
            stock = int(stock_var.get())
        except ValueError:
            return messagebox.showerror("Error", "Invalid price or stock")
        if section not in ["hot", "snacks", "iced_tea", "iced_coffee"]:
            return messagebox.showerror("Error", "Invalid section")
        tree.insert("", "end", values=(section, name, price, stock))
        section_var.set(""), name_var.set(""), price_var.set(""), stock_var.set("")

    def save_changes():
        global hot_coffee_items, snacks_items, iced_tea_items, iced_coffee_items
        hot_coffee_items, snacks_items, iced_tea_items, iced_coffee_items = [], [], [], []

        section_map = {
            "hot": hot_coffee_items,
            "snacks": snacks_items,
            "iced_tea": iced_tea_items,
            "iced_coffee": iced_coffee_items
        }

        for row in tree.get_children():
            section, name, price, stock = tree.item(row)["values"]
            try:
                price = float(price)
                stock = int(stock)
                if section in section_map:
                    section_map[section].append((name, price, stock))
            except ValueError:
                return messagebox.showerror("Error", f"Invalid data for {name}")

        save_menu_to_excel()
        admin_window.withdraw()
        messagebox.showinfo("Saved", "Menu updated successfully.")
        main_window.deiconify()

    def delete_selected():
        selected = tree.selection()
        if not selected:
            return messagebox.showwarning("Warning", "No item selected.")
        if messagebox.askyesno("Confirm Delete", "Are you sure you want to delete this item?"):
            for item in selected:
                tree.delete(item)

    def edit_cell(event):
        row_id = tree.identify_row(event.y)
        col = tree.identify_column(event.x)
        if row_id:
            col_index = int(col.replace("#", "")) - 1
            item = tree.item(row_id)
            new_value = simpledialog.askstring("Edit", f"New value for {item['values'][col_index]}:")
            if new_value is not None:
                item["values"][col_index] = new_value
                tree.item(row_id, values=item["values"])

    tree.bind("<Double-1>", edit_cell)

    # --- Button Frame ---
    button_frame = ttk.Frame(admin_window)
    button_frame.pack(pady=10)

    ttk.Button(button_frame, text="Add Product", command=add_product).grid(row=0, column=0, padx=5)
    ttk.Button(button_frame, text="Save Changes", command=save_changes).grid(row=0, column=1, padx=5)
    ttk.Button(button_frame, text="Delete Selected", command=delete_selected).grid(row=0, column=2, padx=5)

# ------------------- Authentication -------------------

def login():
    username, password = username_entry.get().strip(), password_entry.get().strip()
    users = load_users_from_excel()

    if username == ADMIN_CREDENTIALS["username"] and password == ADMIN_CREDENTIALS["password"]:
        messagebox.showinfo("Login Success", "Welcome, Admin!")
        load_menu_globals()
        open_admin_window()
        main_window.withdraw()
    elif username in users and users[username] == password:
        messagebox.showinfo("Login Success", f"Welcome, {username}!")
        open_post_login_menu(username)
    else:
        messagebox.showerror("Login Failed", "Invalid username or password.")
def open_signup():
    def register():
        new_user, new_pass = new_username.get().strip(), new_password.get().strip()
        users = load_users_from_excel()
        if not new_user or not new_pass:
            return messagebox.showerror("Error", "username and password cannot be empty.")
        if new_user in users or new_user == ADMIN_CREDENTIALS["username"]:
            return messagebox.showerror("Error", "Username already exists or is reserved.")
        add_user_to_excel(new_user, new_pass)
        messagebox.showinfo("Success", "Account created successfully!")
        signup_win.destroy()

    signup_win = tk.Toplevel(main_window)
    signup_win.configure(bg= '#2B1E15')
    signup_win.title("SIGN UP")
    signup_win.resizable(False,False)
    signup_win.geometry('280x130')
    
    
    tk.Label(signup_win, 
             text="New Username:",
             font=('arial',10,'bold'),
             bg='#2B1E15',
             fg='#E0CAA0').grid(row=0,column=0,pady=10,padx=10)
    new_username = tk.Entry(signup_win)
    new_username.grid(row=0,column=1,pady=10,padx=10)
    
    tk.Label(signup_win, text="New Password:",
             font=('arial',10,'bold'),
             bg='#2B1E15',
             fg='#E0CAA0').grid(row=1,column=0,pady=10,padx=10)
    new_password = tk.Entry(signup_win, show="*")
    new_password.grid(row=1,column=1,pady=10,padx=10)
    
    tk.Button(signup_win, 
              text="Register",
              width=14,
              relief= 'flat',
              bg= '#E0CAA0',
              fg= '#2B1E15',
              font=('Arial',10,'bold'),
              command=register).grid(row=3,column=1,pady=5)

# ------------------- Main Window -------------------
ensure_excel_file()
load_menu_globals()

main_window = tk.Tk()
main_window.title("COFFEE LIFE")
main_window.configure(bg='#E0CAA0')
main_window.resizable(False, False)
main_window.geometry("500x500")

tk.Label(main_window,
         text="☕COFFEE LIFE",
         bg='#E0CAA0',
         fg="#2B1E15",
         bd=10,
         font=("Coffee Town",45,'bold')).grid(row=0,
                                              column=0,
                                              columnspan=3,
                                              sticky='ew',
                                              padx=30,
                                              pady=(50,50))

# ---------- user name entry------------
def on_username_focus_in(event):
    if username_entry.get() == 'Username':
        username_entry.delete(0, tk.END)
        username_entry.config(fg='black')
def on_username_focus_out(event):
    if username_entry.get() == '':
        username_entry.insert(0, 'Username')
        username_entry.config(fg='grey')

username_entry = tk.Entry(main_window,
                          font=('arial', 12, 'italic'),
                          fg='grey', width=30, relief='flat')
username_entry.insert(0, 'Username')
username_entry.bind("<FocusIn>", on_username_focus_in)
username_entry.bind("<FocusOut>", on_username_focus_out)
username_entry.grid(row=1, column=0, sticky='n', padx=5, pady=5, ipady=5,columnspan=3)

def on_password_focus_in(event):
    if password_entry.get() == 'Password':
        password_entry.delete(0, tk.END)
        password_entry.config(show='*', fg='black')
def on_password_focus_out(event):
    if password_entry.get() == '':
        password_entry.insert(0, 'Password')
        password_entry.config(show='', fg='grey')

password_entry = tk.Entry(main_window,
                          font=('arial', 12, 'italic'),
                          fg='grey', width=30, relief='flat')
password_entry.insert(0, 'Password')
password_entry.bind("<FocusIn>", on_password_focus_in)
password_entry.bind("<FocusOut>", on_password_focus_out)
password_entry.grid(row=2, column=0, sticky='n', padx=5, pady=(5,30), ipady=5,columnspan=3)

tk.Button(main_window, text=" Sign Up ",
          font=('lets coffee',10,'bold'), 
          bg='#2B1E15',fg='#E0CAA0',
           border=5, relief='flat',
           width= 8, 
           command=open_signup).grid(row=3,column=0,sticky='e')
tk.Button(main_window, text=" Log In ",
          font=('lets coffee',10,'bold'), 
          bg='#2B1E15',fg='#E0CAA0', 
          border=5, relief='flat',
          width= 8, 
          command=login).grid(row=3,column=2,sticky='w')


main_window.grid_columnconfigure(0, weight=1)
main_window.grid_columnconfigure(1, weight=1)
main_window.grid_columnconfigure(2, weight=1)

main_window.mainloop()